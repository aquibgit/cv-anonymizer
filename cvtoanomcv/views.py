

# Create your views here.
import os
from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse
import re
import spacy
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader
import docx

def anonymize_text(text):
    # Simple anonymization (you can improve this)
    text = re.sub(r'\b[\w.-]+?@\w+?\.\w+?\b', '[EMAIL REMOVED]', text)
    text = re.sub(r'\b\d{10}\b', '[PHONE REMOVED]', text)
    return text

def upload_cv(request):
    files_list = []

    if request.method == 'POST' and request.FILES.getlist('cvs'):
        files = request.FILES.getlist('cvs')
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'anonymous'))

        for file in files:
            filename = fs.save(file.name, file)
            file_path = fs.path(filename)

            # If text-based CV, anonymize content
            if filename.endswith('.txt'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                content = anonymize_text(content)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)

            files_list.append({
                'name': filename,
                'url': fs.url(filename)
            })

    return render(request, 'uploadcv.html', {'files': files_list})


nlp = spacy.load("en_core_web_sm")

# -----------------------------------
# Extract text from PDF
# -----------------------------------
def extract_pdf_text(file_path):
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    return text


# -----------------------------------
# Extract text from DOCX
# -----------------------------------
def extract_docx_text(file_path):
    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])


# -----------------------------------
# Advanced Information Extraction
# -----------------------------------
def extract_info(text):

    # -------- EMAIL --------
    email = re.findall(r'\b[\w\.-]+@[\w\.-]+\.\w+\b', text)

    # -------- PHONE --------
    phone = re.findall(r'\+?\d[\d -]{8,12}\d', text)

    # -------- NAME (NLP) --------
    doc = nlp(text[:1000])  # first 1000 chars for speed
    name = "Not Found"
    for ent in doc.ents:
        if ent.label_ == "PERSON":
            name = ent.text
            break

    # -------- SKILLS --------
    skills_master = [
        "Python", "Django", "Machine Learning", "Deep Learning",
        "SQL", "MySQL", "PostgreSQL", "Java", "C++", "React",
        "Angular", "TensorFlow", "Pandas", "NumPy"
    ]

    found_skills = [skill for skill in skills_master if skill.lower() in text.lower()]

    # -------- EDUCATION --------
    education_keywords = ["B.Tech", "B.E", "M.Tech", "MBA", "Bachelor", "Master", "PhD"]
    education_found = [edu for edu in education_keywords if edu.lower() in text.lower()]

    # -------- EXPERIENCE --------
    experience = re.findall(r'(\d+)\+?\s*(years|yrs)', text.lower())
    experience_years = experience[0][0] if experience else "Not Found"

    return {
        "Name": name,
        "Email": email[0] if email else "Not Found",
        "Phone": phone[0] if phone else "Not Found",
        "Skills": ", ".join(found_skills) if found_skills else "Not Found",
        "Education": ", ".join(education_found) if education_found else "Not Found",
        "Experience": experience_years
    }


# -----------------------------------
# Main View
# -----------------------------------
def upload_and_extract(request):

    if request.method == 'POST':

        files = request.FILES.getlist('cvs')
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads'))

        excel_path = os.path.join(settings.MEDIA_ROOT, "cv_database.xlsx")

        # Create Excel if not exists
        if not os.path.exists(excel_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Email", "Phone", "Skills", "Education", "Experience (Years)"])
            workbook.save(excel_path)

        workbook = load_workbook(excel_path)
        sheet = workbook.active

        for file in files:
            filename = fs.save(file.name, file)
            file_path = fs.path(filename)

            if filename.endswith('.pdf'):
                text = extract_pdf_text(file_path)
            elif filename.endswith('.docx'):
                text = extract_docx_text(file_path)
            else:
                continue

            data = extract_info(text)

            sheet.append([
                data["Name"],
                data["Email"],
                data["Phone"],
                data["Skills"],
                data["Education"],
                data["Experience"]
            ])

        workbook.save(excel_path)

        return render(request, 'upload_cv.html', {
            'excel_file': settings.MEDIA_URL + "cv_database.xlsx"
        })

    return render(request, 'uploadcv.html')
